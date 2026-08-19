from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


DEFAULT_EXCEL_PATH = Path(
    r"C:\Users\satya\OneDrive\Desktop\all_output (1).xlsx"
)

# Existing historical dataset already included in this project.
DEFAULT_CSV_PATH = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "processed"
    / "test.csv"
)


def _safe_date(value: Any, year: int, row_index: int) -> datetime:
    if isinstance(value, datetime) and value.year == year:
        return value

    if isinstance(value, date) and value.year == year:
        return datetime.combine(value, datetime.min.time())

    return datetime(year, 1, 1) + timedelta(days=row_index - 1)


def _read_excel_telemetry(workbook_path: Path) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        return []

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if not sheet_name.isdigit():
            continue

        year = int(sheet_name)
        sheet = workbook[sheet_name]

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            generation = (
                row[7]
                if len(row) > 7 and isinstance(row[7], (int, float))
                else None
            )

            if generation is None or generation <= 0:
                continue

            timestamp = _safe_date(
                row[1] if len(row) > 1 else None,
                year,
                row_index,
            )

            insolation = (
                row[11]
                if len(row) > 11 and isinstance(row[11], (int, float))
                else 0
            )

            pr = (
                row[12]
                if len(row) > 12 and isinstance(row[12], (int, float))
                else None
            )

            records.append(
                {
                    "id": f"excel-{sheet_name}-{row_index}",
                    "string_id": "EXCEL_PLANT",
                    "panel_id": "PLANT_TOTAL",
                    "voltage": 1,
                    "current": 1,
                    "power_kw": round(float(generation) / 24, 3),
                    "generation_kwh": round(float(generation), 3),
                    "temperature": 25,
                    "irradiance": round(float(insolation) * 1000 / 24, 3),
                    "performance_ratio": (
                        float(pr) if pr is not None else None
                    ),
                    "timestamp": timestamp.isoformat(),
                }
            )

    return sorted(
        records,
        key=lambda record: record["timestamp"],
    )


def _read_csv_telemetry(csv_path: Path) -> list[dict[str, Any]]:
    if not csv_path.exists():
        return []

    df = pd.read_csv(csv_path)

    required_columns = {
        "Date",
        "Total Generation (KWH)",
        "Insolation",
    }

    if not required_columns.issubset(df.columns):
        return []

    records: list[dict[str, Any]] = []

    for index, row in df.iterrows():
        try:
            timestamp = pd.to_datetime(row["Date"])

            generation = float(row["Total Generation (KWH)"])
            insolation = float(row["Insolation"])

            if generation <= 0:
                continue

            records.append(
                {
                    "id": f"csv-{index + 1}",
                    "string_id": "HISTORICAL_PLANT",
                    "panel_id": "PLANT_TOTAL",
                    "voltage": 1,
                    "current": 1,
                    "power_kw": round(generation / 24, 3),
                    "generation_kwh": round(generation, 3),
                    "temperature": 25,
                    "irradiance": round(insolation * 1000 / 24, 3),
                    "performance_ratio": None,
                    "timestamp": timestamp.isoformat(),
                }
            )

        except (TypeError, ValueError):
            continue

    return sorted(
        records,
        key=lambda record: record["timestamp"],
    )


def read_excel_telemetry(
    path: str | None = None,
) -> list[dict[str, Any]]:
    """
    Read historical telemetry.

    Priority:
    1. Original Excel source, if available.
    2. Existing processed test.csv dataset.

    The API format remains unchanged.
    """

    workbook_path = Path(path or DEFAULT_EXCEL_PATH)

    # First try the original Excel dataset.
    excel_records = _read_excel_telemetry(workbook_path)

    if excel_records:
        return excel_records

    # Fallback to the historical CSV already included in this project.
    return _read_csv_telemetry(DEFAULT_CSV_PATH)